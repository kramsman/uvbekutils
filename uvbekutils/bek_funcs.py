""" small functions that can be imported and used in other programs.

Use uv to manage environments.  Use uv add git+https://github.com/kramsman/bekutils.git
# maybe: uv add uvbekutils --upgrade-package uvbekutils

To run intalling package from within code:
import subprocess
subprocess.run(["uv", "add", "git+https://github.com/kramsman/uvbekutils.git"])
# uv add uvbekutils --upgrade-package uvbekutils

NO NORE CONDA
To add to an environment, in terminal:
    conda activate ROVGeneral
    pip install --force-reinstall git+https://github.com/kramsman/bekutils.git

"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# TODO what to do with loggers?

log_level = "DEBUG"  # used for log file; screen set to INFO. TRACE, DEBUG, INFO, WARNING, ERROR

def safe_str(value: object) -> str:
    """Convert a value to string, returning empty string for NaN.

    Args:
        value: Any value to convert; NaN yields empty string.

    Returns:
        Empty string if value is NaN, otherwise str(value).
    """
    import pandas as pd
    return '' if pd.isna(value) else str(value)


def load_workbook_w_filepath(file: Path, *args, **kwargs) -> Workbook:
    """Load an openpyxl workbook and attach the source path as a filepath attribute.

    Args:
        file: Path to the workbook file.
        *args: Positional arguments passed to openpyxl.load_workbook.
        **kwargs: Keyword arguments passed to openpyxl.load_workbook.

    Returns:
        The loaded Workbook with a filepath attribute set to file.
    """
    from openpyxl import load_workbook
    wb = load_workbook(file, *args, **kwargs)
    wb.filepath = file
    return wb


def wb_path(wb: Workbook) -> Path | None:
    """Return the filepath attribute of a workbook, or None if not set.

    Can reference book name of a worksheet via ws.parent.filepath.name.

    Args:
        wb: An openpyxl Workbook, typically loaded with load_workbook_w_filepath.

    Returns:
        The Path stored in wb.filepath, or None if the attribute does not exist.
    """

    # return (wb.filepath if hasattr(wb,'filepath') else None)
    if hasattr(wb, 'filepath'):
        return wb.filepath
    else:
        return None


def wb_name(wb: Workbook) -> str:
    """Return the filename of a workbook as a string, or 'Unknown' if not set.

    Args:
        wb: An openpyxl Workbook, typically loaded with load_workbook_w_filepath.

    Returns:
        The filename string from wb.filepath.name, or 'Unknown' if the
        filepath attribute does not exist.
    """

    if hasattr(wb, 'filepath'):
        return wb.filepath.name
    else:
        return 'Unknown'


def bek_excel_titles(
    wb: Workbook,
    sheet_name_list: list | str,
    cell_infos: list | None = None,
    auto_size_before: bool | None = None,
    auto_size_after: bool | None = None,
) -> None:
    """Write formatted cell values and titles to sheets in an Excel workbook.

    Autosizing before title insertion is recommended to avoid '######' in cells,
    since title text is typically wider than column data.

    Args:
        wb: The openpyxl Workbook to modify.
        sheet_name_list: Sheet name or list of sheet names to add titles to.
            Pass True to apply to all sheets.
        cell_infos: List of cell-attribute dicts. Each dict specifies a cell
            by 1-based row/col and a cell attribute to set. Use 'value' for
            text and 'font' for a Font(...) expression string. Example::

                [
                    {'row': 1, 'col': 1, 'cell_attr': 'value', 'cell_value': 'Report'},
                    {'row': 1, 'col': 1, 'cell_attr': 'font',  'cell_value': 'Font(b=True, size=20)'},
                ]

        auto_size_before: If truthy, autosize all columns before writing cells.
        auto_size_after: If truthy, autosize all columns after writing cells.
    """

    from openpyxl.styles import Font
    from uvbekutils import autosize_xls_cols

    if isinstance(sheet_name_list,str):
        sheet_name_list = [sheet_name_list]
    elif isinstance(sheet_name_list, list):
        pass
    else:
        exit_yes(f"sheet_name_list is not str or list: {sheet_name_list=}", "Error", raise_err=True)

    if auto_size_before:
        for sh in wb.worksheets:
            autosize_xls_cols(sh)

    # TODO perform checks on formats of cell_infos
    if cell_infos:
        for sh in wb.worksheets:
            if sh.title in sheet_name_list:
                for cell_info in cell_infos:
                    if cell_info['cell_attr'] == 'font':
                        # use eval because needs to be like ft1 = Font(name='Arial', size=14).  might be able to use
                        # another setattr but not ready to try now
                        setattr(sh.cell(row=cell_info['row'], column=cell_info['col']), cell_info['cell_attr'],
                                eval(cell_info['cell_value']))
                    else:
                        setattr(sh.cell(row=cell_info['row'], column=cell_info['col']), cell_info['cell_attr'],
                                cell_info['cell_value'])

    if auto_size_after:
        for sh in wb.worksheets:
            autosize_xls_cols(sh)


def bek_write_excel(
    df: pd.DataFrame,
    sheet_name: str,
    startrow: int,
    cell_infos: list | None = None,
) -> None:
    """Write a DataFrame to an Excel file named after the running script.

    Autosizes columns before writing title cells to avoid '######'. The output
    file is placed in the same directory as the running executable.

    Args:
        df: DataFrame to write to Excel.
        sheet_name: Name of the worksheet to write the DataFrame to.
        startrow: Row index (0-based) where the DataFrame header will be written.
        cell_infos: Optional list of cell-attribute dicts for formatting. See
            bek_excel_titles for format details.
    """

    from pathlib import Path
    import pandas as pd
    from openpyxl.styles import Font
    from uvbekutils import autosize_xls_cols
    from uvbekutils import exe_file

    op_file = exe_file().with_suffix(".xlsx")

    writer = pd.ExcelWriter(op_file)

    df.to_excel(writer, sheet_name = sheet_name, startrow = startrow)
    wb = writer.book
    for sh in wb.worksheets:
        autosize_xls_cols(sh)

    if cell_infos:
        for sh in wb.worksheets:
            for cell_info in cell_infos:
                if cell_info['cell_attr'] == 'font':
                    # use eval because needs to be like ft1 = Font(name='Arial', size=14).  might be able to use
                    # another setattr but not ready to try now
                    setattr(sh.cell(row=cell_info['row'], column=cell_info['col']), cell_info['cell_attr'],
                            eval(cell_info['cell_value']))
                else:
                    setattr(sh.cell(row=cell_info['row'], column=cell_info['col']), cell_info['cell_attr'],
                            cell_info['cell_value'])

    writer.close()

def exe_file() -> Path:
    """Return the Path of the currently running script or frozen executable.

    Handles three cases: frozen executable (PyInstaller), normal script run,
    and fallback to sys.argv[0] or 'app'.

    Returns:
        Path to the running executable or script file.
    """

    import sys
    from pathlib import Path

    import __main__

    # determine if application is running as a script file or frozen exe
    if getattr(sys, 'frozen', False):
        exe_file = Path(sys.executable)
    # elif __file__:
    #     # exe_file = Path(__file__).parents[0]
    #     exe_file = Path(__main__.__file__)
    elif hasattr(__main__, '__file__') and __main__.__file__:
        exe_file = Path(__main__.__file__)
    elif sys.argv[0]:
        exe_file = Path(sys.argv[0])
    else:
        exe_file = Path("app")

    return exe_file

def exe_path() -> Path:
    """Return the directory Path containing the currently running executable.

    Returns:
        Parent directory of exe_file(), or the current working directory
        if exe_file() returns None.
    """

    from pathlib import Path
    from uvbekutils import exe_file
    result = exe_file()
    if result is None:
        exe_path = Path.cwd()
    else:
        exe_path = result.parents[0]

    return exe_path


def setup_loguru(
    log_level_std: str = 'INFO',
    log_level_log: str = 'INFO',
    log_path: Path | None = None,
    log_mode: str = 'w',
    log_file: bool = True,
) -> logger:
    """Configure and return a loguru logger with stdout and optional file sinks.

    Log file is placed in the same directory as the running executable. Any
    existing log file is removed before a new one is opened.

    Args:
        log_level_std: Minimum log level for stdout output (e.g. 'INFO', 'DEBUG').
        log_level_log: Minimum log level for the log file (e.g. 'DEBUG', 'TRACE').
        log_path: Directory for the log file. Defaults to exe_path() if None.
        log_mode: File open mode for the log file ('w' to overwrite, 'a' to append).
        log_file: If True, create a log file sink in addition to stdout.

    Returns:
        The configured loguru logger instance.
    """

    # TODO suppress error if log not found which happens on first run
    # TODO do not create lof if log level is blank

    from pathlib import Path
    from loguru import logger
    import os
    import sys
    from uvbekutils import exe_path
    from uvbekutils import exe_file

    # LOG_LEVEL_LOG = "TRACE"  # used for log file; screen set to INFO. TRACE, DEBUG, INFO, WARNING, ERROR
    # LOG_LEVEL_STD = "DEBUG"  # used for log file; screen set to INFO. TRACE, DEBUG, INFO, WARNING, ERROR

    logger.trace("entered setup_loguru")
    if log_path is None:
        logger.trace("setting log_path if non w exe_path()")
        log_path = exe_path()

        # # determine if application is running as a script file or frozen exe
        # if getattr(sys, 'frozen', False):
        #     log_path = Path(sys.executable).parents[0]
        # elif __file__:
        #     # root_path = os.path.dirname(__file__)
        #     log_path = Path(__file__).parents[0]
        # else:
        #     log_path = None
    logger.debug(f"({log_path=}")

    logger.trace("setting sys.stdout")
    logger.add(sys.stdout, level=log_level_std, backtrace=True, diagnose=False)

    if log_file:
        logger.trace("setting log_file info - logger.remove(0) next")

        try:
            logger.remove(0)
        except ValueError:
            pass

        logfile = exe_file().with_suffix(".log")
        try:
            os.remove(logfile)
        except Exception as e:
            logger.exception(e)
            pass

        logger.trace("setting log_file info - adding logfile")
        logger.add(open(logfile, log_mode), level=log_level_log, backtrace=True, diagnose=False)

    return logger


def exit_yes_no(msg: str, title: str | None = None, display_exiting: bool = False) -> None:
    """Display a Yes/No prompt and exit the program if the user chooses No.

    Args:
        msg: Message text to display in the dialog.
        title: Optional dialog window title.
        display_exiting: If True, show an 'Exiting' alert before exiting.
    """

    from uvbekutils import pyautobek
    from loguru import logger

    choice = pyautobek.confirm(msg, title, buttons=['Continue', 'Exit'])
    if choice == "exit":
        if display_exiting:
            pyautobek.alert("Exiting", "Alert")
        logger.debug("here")
        exit()


def exit_yes(msg: str, title: str | None = None, *, errmsg: str | None = None, raise_err: bool = False) -> None:
    """Display an alert popup and then exit or raise an exception.

    Args:
        msg: Message text displayed in the popup.
        title: Optional dialog window title. Defaults to '** Exiting Program **'.
        errmsg: Error message for the raised exception. Defaults to msg with
            newlines removed.
        raise_err: If True, raise an Exception instead of calling exit().
    """

    from uvbekutils import pyautobek
    from loguru import logger

    if not errmsg:
        errmsg = msg.replace("\n", " ")  # dont fill the console with linefeeds
    if not title:
        title = "** Exiting Program **"
    logger.debug("in 'exit_yes'")
    pyautobek.alert(msg, title)
    if raise_err:
        logger.debug("ready to raise error'")
        raise Exception(errmsg)
    else:
        logger.debug("ready to exit'")
        exit()


def is_number(s: str) -> bool:
    """Check whether a value can be interpreted as a number.

    Handles np.nan explicitly, since np.nan is a float and would otherwise
    pass a float() conversion check.

    Args:
        s: Value to test, expected to be passed as a string.

    Returns:
        True if s can be converted to float and is not NaN, False otherwise.
    """

    import numpy as np
    from loguru import logger

    logger.trace("in is_number")

    try:
        if np.isnan(s):  # this is needed- np.nan are int which are numbers
        # if s == np.nan:  # this is needed- np.nan are int which are numbers
            return False
    except TypeError:
        pass
    if s is None:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def clean_field(fld: str, case_convert: str = 'lower') -> str:
    """Normalize a string by stripping whitespace and removing special characters.

    Removes spaces, apostrophes, periods, and hyphens. Optionally converts case.
    Compatible with DataFrame.apply: ``df['col'].apply(clean_field)``.

    Args:
        fld: Value to clean; will be cast to str before processing.
        case_convert: Case conversion to apply. One of 'lower', 'upper', or
            'keep' (no conversion).

    Returns:
        Cleaned and case-converted string.
    """

    #TODO: pass characters to be removed as string

    from loguru import logger

    return_fld = str(fld).strip().replace(" ", "").replace("'", "").replace(".", "").replace("-", "")
    if case_convert == 'lower':
        return_fld = return_fld.lower()
    elif case_convert == 'upper':
        return_fld = return_fld.upper()
    elif case_convert == 'keep':
        pass
    else:
        exit_yes(f"wrong value fed to clean_field parameter case_convert, '{case_convert}' - exiting")
    return return_fld


def autosize_xls_cols(ws: Worksheet) -> None:
    """Auto-fit column widths in an openpyxl worksheet to their content.

    Datetime cells are treated as width 10. All other cells use the string
    length of their value.

    Args:
        ws: The openpyxl Worksheet to resize.
    """

    from loguru import logger

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                if cell.data_type == 'd':
                    date_width = 10
                else:
                    date_width = len(str(cell.value))
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), date_width))

    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 1


def bad_file_exit(file: Path, msg: str | None = None, raise_err: bool = False) -> None:
    """Exit or raise an error if the given file does not exist.

    Args:
        file: Path to the file to check.
        msg: Custom message for the alert. Defaults to a standard
            'file does not exist' message.
        raise_err: If True, raise an Exception instead of calling exit().
    """

    from loguru import logger

    if msg is None:
        msg = f"File:\n\n'{file}'\n\ndoes not exist."
    if not file.expanduser().exists():
        logger.debug("here")
        exit_yes(msg, raise_err=raise_err)


def bad_path_exit(path: Path, msg: str | None = None, raise_err: bool = False) -> None:
    """Exit or raise an error if the given directory does not exist.

    Args:
        path: Path to the directory to check.
        msg: Custom message for the alert. Defaults to a standard
            'directory does not exist' message.
        raise_err: If True, raise an Exception instead of calling exit().
    """

    from loguru import logger

    if msg is None:
        msg = f"Directory:\n\n'{path}'\n\ndoes not exist."
    if not path.expanduser().exists():
        logger.debug("here")
        exit_yes(msg, raise_err=raise_err)


def bad_path_create(path: Path, msg: str | None = None) -> None:
    """Create a directory if it does not exist, alerting the user first.

    Args:
        path: Path to the directory to create if missing.
        msg: Custom alert message. Defaults to a standard message naming
            the path and the calling function.
    """

    import os
    from uvbekutils import pyautobek
    from loguru import logger

    if msg is None:
        msg = ("Directory:\n\n" + str(path) + "\n\ndoes not exist.  Creating." +
               "\n\nCalled from " + calling_func(level=2))
    if not os.path.isdir(path):
        logger.debug("here")
        pyautobek.alert(msg, "Adding Directory via bad_path_create")
        os.makedirs(path)


def calling_func(level: int = 0) -> str:
    """Return the name and line number of a function in the call stack.

    Args:
        level: Stack depth to inspect. 0 is this function, 1 is its caller,
            2 is the caller's caller, etc.

    Returns:
        A string with the function name and line number, or an error
        description if the stack level is too deep.
    """

    import inspect
    from loguru import logger

    try:
        func = f"'{inspect.stack()[level][3]}', line #: {inspect.stack()[level][2]}"
    except Exception as e:
        logger.exception(e)
        func = f"** error ** inspect level too deep: {str(level)} called from {inspect.stack()[level][3]}"
    return func


def read_file_to_df(file_with_path: Path, **param_dict) -> pd.DataFrame | None:
    """Read an xlsx or csv file into a DataFrame, filtering kwargs by file type.

    Only keyword arguments valid for pd.read_excel or pd.read_csv are passed
    through; unsupported keys are silently dropped.

    Args:
        file_with_path: Path to the input file (.xlsx or .csv).
        **param_dict: Optional keyword arguments forwarded to the appropriate
            pandas read function.

    Returns:
        DataFrame with the file contents, or None if the file type is not
        supported (in practice exits via exit_yes before returning None).
    """

    import inspect
    from pathlib import Path
    import pandas as pd
    from loguru import logger

    logger.info(f"reading file to dataframe '{file_with_path.stem}'")

    if Path(file_with_path).suffix.lower() == '.xlsx':
        if param_dict is not None:
            filtered_dict = {k: v for k, v in param_dict.items()
                             if k in [p.name for p in inspect.signature(pd.read_excel).parameters.values()]}
        else:
            filtered_dict = {}
        df_temp = pd.read_excel(file_with_path, **filtered_dict)
    elif file_with_path.suffix.lower() == '.csv':
        filtered_dict = {k: v for k, v in param_dict.items()
                         if k in [p.name for p in inspect.signature(pd.read_csv).parameters.values()]}
        df_temp = pd.read_csv(file_with_path, **filtered_dict)
    else:
        df_temp = None
        logger.debug('here')
        exit_yes((f"Bad file type on input - not xlsx or csv."
                  f"\n\nFile: '{file_with_path}'"
                  ))
    return df_temp


def find_header_row_in_file(
    file_with_path: Path,
    header_string: str,
    header_col: str,
    sheet_name: str | int | None = None,
) -> int:
    """Find the row index of a header by matching a string in a specific column.

    Searches only the first 30 rows. Useful for files that have title or blank
    rows above the actual data header.

    Args:
        file_with_path: Path to the input file (.xlsx or .csv).
        header_string: String expected in the header cell (case-insensitive).
        header_col: Excel-style column letter(s) to search in (e.g. 'A', 'B', 'AA').
        sheet_name: Sheet name or index. Defaults to the first sheet (0) if None.

    Returns:
        0-based row index of the header row.
    """

    from openpyxl.utils.cell import coordinate_from_string
    from openpyxl.utils.cell import column_index_from_string
    from uvbekutils import read_file_to_df

    if sheet_name is None:
        sheet_name = 0
    header_row = None

    # If the header identifying field is not in the first 30 rows assume something is wrong in the file
    df_temp = read_file_to_df(file_with_path, **{'header': None, 'sheet_name': sheet_name, 'nrows': 30,
                                                 'keep_default_na': True, 'dtype': str})

    excel_col_num = column_index_from_string(coordinate_from_string(header_col + '1')[0])  # -1 to 0 index

    for row in df_temp.itertuples():
        if type(row[excel_col_num]) == str:  # cell in input being checked is ok, otherwise blank cells/None cause
            # problems in compare
            if row[excel_col_num].strip().lower() == header_string.strip().lower():
                header_row = row.Index
                break
    if header_row is None:
        exit_yes((f"File may be bad.\n\nThe header check string '{header_string}' "
                  f"was not found in column '{header_col}' "
                  "in the first 30 lines of input file:"
                  f"\n\n'{file_with_path}'"
                  ))
    return header_row


def check_ws_headers(ws: Worksheet, vals: list[tuple[str, str]]) -> None:
    """Verify that worksheet cells match expected header label strings.

    Args:
        ws: The openpyxl Worksheet to check.
        vals: List of (cell_address, expected_value) tuples, e.g.
            [('A1', 'use'), ('B1', 'fromFilePath')].
    """

    def chk_header_vals(ws_to_chk: Worksheet, cell: str, val: str) -> None:
        """Exit with an error if the cell value does not match val.

        Args:
            ws_to_chk: Worksheet containing the cell to check.
            cell: Cell address string (e.g. 'A1').
            val: Expected cell value (case-insensitive comparison).
        """
        if str(ws_to_chk[cell].value).strip().lower() != str(val).lower():
            exit_yes((f"Column heading '{cell}' on Setup sheet '{ws_to_chk.title}' not equal to literal '{val}'."
                      f"\n\nIt is '{str(ws_to_chk[cell].value)}'."),
                     )

    for pairs in vals:
        chk_header_vals(ws, pairs[0], pairs[1])

# # TODO Add in check_fie+headers like above with csv
#
# # def text_box(txt: str, title: str = '', box_title: str = '', buttons: list | None = None) -> str | None:
# #     """Display a scrollable text block in a PySimpleGUI window with custom buttons.
# #
# #     Window size is auto-scaled to the text content, capped at 80 columns and
# #     80 rows. A vertical scrollbar is added when the text exceeds 80 lines.
# #
# #     Args:
# #         txt: Text to display, with lines separated by '\\n'.
# #         title: Heading text rendered inside the window above the text area.
# #         box_title: Title shown in the window title bar.
# #         buttons: List of button label strings. Defaults to ['OK', 'Exit'].
# #
# #     Returns:
# #         Lowercased label of the button clicked, or None if the window was
# #         closed without a button press.
# #     """
# #
# #     import PySimpleGUI as sg
# #
# #     # window = sg.Window('Virus Simulation', layout, background_color='hex_color_code')
# #
# #     if buttons is None:
# #         buttons = ["OK", "Exit"]
# #
# #     col_factor = 3  # to scale window equally
# #     row_factor = 30  # to scale window equally
# #     max_cols = len(max(txt.split("\n"), key=len)) * col_factor
# #     cols = max_cols
# #     # v_scroll = False
# #     col_limit = 80 * col_factor
# #     col_min = 50 * col_factor
# #     if cols > col_limit:
# #         # v_scroll = True
# #         cols = col_limit
# #     elif cols < col_min:
# #         cols = col_min
# #
# #     noscroll = True
# #     row_limit = 80
# #     row_min = 6
# #     # max_rows = len(txt.split("\n"))
# #     # rows = max_rows
# #     rows = len(txt.split("\n"))
# #     if rows > row_limit:
# #         noscroll = False
# #         rows = row_limit
# #     elif rows < row_min:
# #         rows = row_min
# #     #horizontal_scroll=h_scroll,
# #     # sg.theme('SystemDefault1')
# #     sg.theme('Default1')
# #     layout = [
# #         [sg.Text(title, font=("Arial", 18))],
# #         [sg.Multiline(txt, autoscroll=False, expand_x=True, no_scrollbar=noscroll,
# #                       expand_y=True, enable_events=True)],
# #         [sg.Button(text) for text in buttons],
# #     ]
# #
# #     event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
# #                               use_custom_titlebar=True, size=(600, rows*row_factor), disable_close=True,
# #                               resizable=True, grab_anywhere=True).read(close=True)
# #     if event is not None:
# #         event = event.lower()
# #     return event
#
#
# # def get_dir_name(box_title: str, title2: str, initial_dir: str | Path) -> Path:
# #     """Show a folder-picker dialog and return the selected directory as a Path.
# #
# #     Exits via exit_yes if no directory is chosen.
# #
# #     Args:
# #         box_title: Title shown in the window title bar.
# #         title2: Heading text displayed inside the dialog.
# #         initial_dir: Starting directory for the folder browser.
# #
# #     Returns:
# #         Expanded Path of the chosen directory.
# #     """
# #
# #     import os
# #     import PySimpleGUI as sg
# #     from pathlib import Path
# #     from loguru import logger
# #
# #     logger.debug('in get_dir_name')
# #
# #     layout = [
# #         [sg.Text(title2, font=("Arial", 18))],
# #         [
# #          sg.Input(key="-IN-", expand_x=True),
# #          sg.FolderBrowse(initial_folder=Path(initial_dir).expanduser())
# #          ],
# #         [sg.Button("Choose")],
# #     ]
# #
# #     # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
# #     event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
# #                               size=(1000, 150), use_custom_titlebar=True).read(close=True)
# #
# #     dir_name = values['-IN-']
# #     if dir_name == "":
# #         exit_yes("No directory name chosen")
# #
# #     return Path(dir_name).expanduser()
#
#
# def get_file_name(box_title: str, title2: str, initial_dir: str | Path) -> Path:
#     """Show a file-picker dialog and return the selected file as a Path.
#
#     Exits via exit_yes if no file is chosen.
#
#     Args:
#         box_title: Title shown in the window title bar.
#         title2: Heading text displayed inside the dialog.
#         initial_dir: Starting directory for the file browser.
#
#     Returns:
#         Expanded Path of the chosen file.
#     """
#
#     import PySimpleGUI as sg
#     from pathlib import Path
#     from loguru import logger
#     from uvbekutils import exit_yes
#
#     logger.debug('in get_file_name')
#     # "Select Sincere address export file 'all-parent-campaign-requests-yyyy-mm-dd.csv'"
#     layout = [
#         [sg.Text(title2, font=("Arial", 18))],
#         [
#          sg.Input(key="-IN-", expand_x=True),
#          sg.FileBrowse(initial_folder=Path(initial_dir).expanduser())
#          ],
#         [sg.Button("Choose")],
#     ]
#
#     # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
#     event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
#                               size=(1000, 200), use_custom_titlebar=True).read(close=True)
#     # sg.Window.close()
#
#     file_name = values['-IN-']
#     if file_name == "":
#         exit_yes("No file name chosen")
#
#     return Path(file_name).expanduser()


def convert_bool(bool_val: bool | str | None) -> bool:
    """Convert a bool or string representation to a Python bool.

    Handles the limitation that bool('FALSE') == True in Python.

    Args:
        bool_val: A bool, or a string of any case matching 'true' or 'false'.

    Returns:
        True or False.

    Raises:
        ValueError: If bool_val is None or not a recognized boolean string.
    """
    if isinstance(bool_val, bool):
        return_val = bool_val
    else:
        if bool_val is None or bool_val.lower() not in ['true', 'false']:
            raise ValueError('only allowable booleans are any case of true and false.  0/1 could be added to '
                             'convert_bool code')
        elif bool_val.lower() == 'true':
            return_val = True
        else:
            return_val = False
    return return_val


def conc_addr(concentration_dict: dict, state: str | None = None, city: str | None = None, address: str | None = None) -> bool:
    """Check whether a state/city/address combination is in the concentration dictionary.

    State, city, and address are cleaned using the 'clean_field' function
    before lookup.

    Args:
        concentration_dict: Dictionary mapping (state, city, address) tuples
            to concentration metadata.
        state: State code or name.
        city: City name.
        address: Street address.

    Returns:
        True if the cleaned address is found in concentration_dict,
        False otherwise.
    """

    from uvbekutils import clean_field

    concentrated = (True if (clean_field(state), clean_field(city), clean_field(address)) in
                             concentration_dict else False)
    return concentrated


def conc_addr_desc(concentration_dict: dict, state: str | None = None, city: str | None = None, address: str | None = None) -> str:
    """Return the description for a concentrated address entry.

    State, city, and address are cleaned using the 'clean_field' function
    before lookup.

    Args:
        concentration_dict: Dictionary mapping (state, city, address) tuples
            to concentration metadata.
        state: State code or name.
        city: City name.
        address: Street address.

    Returns:
        The 'desc' string for the matching address entry, or an empty string
        if not found in concentration_dict.
    """

    from uvbekutils import clean_field

    desc = concentration_dict.get((clean_field(state), clean_field(city),
                                               clean_field(address)), {'desc': "", 'remove': ""})['desc']
    return desc

def conc_addr_remove_desc(concentration_dict: dict, state: str | None = None, city: str | None = None, address: str | None = None) -> str:
    """Return the removal description for a concentrated address entry.

    State, city, and address are cleaned using the 'clean_field' function
    before lookup.

    Args:
        concentration_dict: Dictionary mapping (state, city, address) tuples
            to concentration metadata.
        state: State code or name.
        city: City name.
        address: Street address.

    Returns:
        The 'remove' string for the matching address entry, or an empty string
        if not found in concentration_dict.
    """

    from uvbekutils import clean_field

    desc = concentration_dict.get((clean_field(state), clean_field(city),
                                               clean_field(address)), {'desc': "", 'remove': ""})['remove']
    return desc


def scroll_box(txt: str, *, title: str | None = None, wrap_lines: bool = True) -> None:
  """Display a read-only scrollable text box using a PySide6 Qt window.

  Reuses an existing QApplication if one is already running, otherwise
  creates a new one and blocks until the window is closed.

  Args:
      txt: Text content to display.
      title: Window title bar text.
      wrap_lines: If True, wrap long lines to the window width. If False,
          enable a horizontal scrollbar instead.
  """

  from PySide6.QtWidgets import QApplication, QMainWindow, QTextEdit
  import sys

  # Reuse existing QApplication if one exists, otherwise create new one
  app = QApplication.instance()
  app_created = False
  if app is None:
      app = QApplication(sys.argv)
      app_created = True

  # Create main window with title
  window = QMainWindow()
  window.setWindowTitle(title)

  # Create text edit widget
  text_edit = QTextEdit()
  text_edit.setPlainText(txt)
  text_edit.setReadOnly(True)

  # Enable or disable line wrapping which introduces horizontal scroll bar
  if wrap_lines:
      text_edit.setLineWrapMode(QTextEdit.WidgetWidth)
  else:
      text_edit.setLineWrapMode(QTextEdit.NoWrap)

  # Set text edit as the central widget
  window.setCentralWidget(text_edit)
  window.resize(600, 400)
  window.show()

  if app_created:
      sys.exit(app.exec())
  else:
      app.exec()



if __name__ == '__main__':

    txt = 'This is the first line. \n\n' + " random text"*100
    scroll_box(txt, wrap_lines=False, title="box")
    # exit_yes_no("this is the msg", "Title", display_exiting=True)
    # exit_yes("this is the msg", "Title", errmsg="Err msg", raise_err = False)
    a=1
